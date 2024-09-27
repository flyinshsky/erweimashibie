import cv2   
from pyzbar.pyzbar import decode  
import numpy as np  
from openpyxl import Workbook  
import time  
import logging  
from datetime import datetime  # 导入datetime模块来获取时间戳
  
# 初始化 Excel 工作簿和工作表  
wb = Workbook()  
ws = wb.active  
ws.title = "QRCodeData"  

# 初始化窗口并设置初始大小
def init_window(window_name, width, height):
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)  # 允许调整窗口大小
    cv2.resizeWindow(window_name, width, height)  # 设置初始窗口大小
  
# 设置自动保存间隔（秒）  
AUTO_SAVE_INTERVAL = 10  
  
# 用来跟踪是否应该暂停自动保存的变量  
auto_save_paused = False  
  
# log information settings  
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')  
  
def decode_and_display(image, ws, row_index):  
    barcodes = decode(image)  
    for barcode in barcodes:  
        barcode_data = barcode.data.decode("utf-8")  
        barcode_type = barcode.type  
        # 写入到 Excel  
        ws.cell(row=row_index, column=1, value=barcode_data) 
        ws.cell(row=row_index, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])  # 写入时间戳，毫秒保留三位 
        row_index += 1  
        # 绘制矩形和文本显示（如果需要）  
        (x, y, w, h) = barcode.rect  
        cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)  
        text = "{} ({})".format(barcode_data, barcode_type)  
        cv2.putText(image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)  
    return image, row_index  
  
def save_excel(filename):  
    wb.save(filename)  
    logging.info(f"Excel 文件 {filename} 已保存。")  
  
def save_image(img, filename):  
    cv2.imwrite(filename, img)  
    logging.info(f"图像文件 {filename} 已保存。")  
  
def detect_and_save():  
    cap = cv2.VideoCapture(0)  
    row_index = 1  # Excel 中的起始行  
    last_save_time = time.time()  # 记录上次保存的时间  

    # 初始化窗口并设置大小
    init_window("QR Code Scanner", 1024, 600)  # 假设初始窗口大小为 1024x600
    
    while True:  
        ret, img = cap.read()  
        if not ret:  
            break  
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  
        image, row_index = decode_and_display(gray, ws, row_index)  
        current_time = time.time()  
          
        if current_time - last_save_time >= AUTO_SAVE_INTERVAL and not auto_save_paused:  
            # 保存 Excel 文件  
            save_excel("QRCodeData_auto.xlsx")  
            # （如果需要）保存带有 QR 码的图像  
            save_image(img, "QRCodeData_auto_image.png")  # 这里可以根据需要取消注释  
            last_save_time = current_time  
  
        # 如果检测到 QR 码，则保存图像（这里我们总是保存，不论是否自动保存）  
        if len(decode(gray)) > 0:  
            save_image(img, "qrcode_image.png")  
  
        # 显示图像  
        cv2.imshow('QR Code Scanner', image)  
  
        k = cv2.waitKey(1000)  
        if k == ord('s'):  # 手动保存 Excel  
            save_excel("QRCodeData_manual.xlsx")  
        elif k == ord('i'):  # 手动保存图像  
            save_image(img, "qrcode_image_manual.png")  
        elif k == 27:  # 退出  
            break  
  
    cap.release()  
    cv2.destroyAllWindows()  
  
# 开始检测和保存  
detect_and_save()

# 代码现在已经整合了二维码解码、Excel数据写入、图像显示以及
# 通过特定按键（ESC退出摄像，'s'手动保存）来触发保存的功能。
# 在这个代码中，detect_and_save 函数负责打开摄像头、解码二维码、
# 在Excel中写入数据、显示图像，并根据用户的按键输入决定是否保存Excel文件和图像。
# 如果按下ESC键，程序将退出摄像；如果按下's'键，程序将保存Excel文件和当前摄像头的图像。
# 注意，图像保存的是原始的彩色图像，而不是处理后的灰度图像。
# 在上面的代码中，会每过10秒自动保存一次表格和图片，我添加了save_excel和save_image函数来分别保存Excel和图像文件。在detect_and_save函数中，我添加了键盘监听来允许用户手动保存Excel（按's'键）和图像（按'i'键）。
# Excel文件的自动保存间隔。当这个间隔过去后，并且auto_save_paused为False时，程序会自动保存Excel文件（这里并没有保存带有QR码的图像作为默认行为，但你可以取消上面代码中关于图像保存的注释行来实现）。detect_and_save函数中使用了一个无限循环来不断地从摄像头读取图像，解码QR码，并在屏幕上显示结果。我使用了logging模块来记录保存文件的信息，这样你可以看到何时何地保存了文件。Excel文件的自动保存间隔。当这个间隔过去后，并且auto_save_paused为False时，程序会自动保存Excel文件（这里并没有保存带有QR码的图像作为默认行为，但你可以取消上面代码中关于图像保存的注释行来实现）。detect_and_save函数中使用了一个无限循环来不断地从摄像头读取图像，解码QR码，并在屏幕上显示结果。我使用了logging模块来记录保存文件的信息，这样你可以看到何时何地保存了文件。